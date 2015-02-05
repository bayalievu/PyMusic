#Processing Music Database from xlsx file using Python
#comment
from openpyxl import load_workbook
wb = load_workbook("test.xlsx")

#for s in wb.get_sheet_names():
#sheet = wb.get_sheet_by_name("")
sheet = wb.worksheets[3]
for row in sheet.iter_rows(row_offset=1):
        f = []
        a = []
        song = ""
        if row[0] is not None and row[0].value is not None:
                f.append(row[0].value.strip())
                a.append(row[0].value.strip())
        if row[1] is not None and row[1].value is not None:     
                f.append(row[1].value.strip())
                a.append(row[1].value.strip())
        if row[2] is not None and row[2].value is not None:     
                f.append(row[2].value.strip())
                song = row[2].value.strip()
        filename = "_".join(f)
        artist = " ".join(a)
        print "Artist: " + artist
        print "Song: " + song
        print "File: "+ filename + ".mp3"


